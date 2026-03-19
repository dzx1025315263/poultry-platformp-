#!/usr/bin/env python3
"""
修复全球禽肉进口商数据库 v2：
1. 修正国家/地区归类错误（根据企业简介判断真实国家）
2. 规范化所有文本字段：英文在前，中文在后
3. 输出新的 Excel 文件
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import json

INPUT_FILE = '/home/ubuntu/upload/全球禽肉进口商数据库_全新版.xlsx'
OUTPUT_FILE = '/home/ubuntu/全球禽肉进口商数据库_修正版.xlsx'

# ============================================================
# 国家 -> 大洲 映射
# ============================================================
COUNTRY_CONTINENT = {
    # 中东
    '伊拉克': '中东', '阿联酋': '中东', '卡塔尔': '中东', '沙特阿拉伯': '中东',
    '阿曼': '中东', '科威特': '中东', '巴林': '中东', '约旦': '中东',
    '伊朗': '中东', '叙利亚': '中东', '也门': '中东', '黎巴嫩': '中东',
    '土耳其': '中东',
    # 非洲
    '吉布提': '非洲', '苏丹': '非洲', '突尼斯': '非洲', '科特迪瓦': '非洲',
    '埃塞俄比亚': '非洲', '埃及': '非洲', '摩洛哥': '非洲', '加纳': '非洲',
    '利比亚': '非洲', '尼日利亚': '非洲', '南非': '非洲', '安哥拉': '非洲',
    '利比里亚': '非洲', '塞拉利昂': '非洲', '几内亚': '非洲', '贝宁': '非洲',
    '中非': '非洲', '中非共和国': '非洲', '刚果（金）': '非洲', '刚果民主共和国': '非洲',
    '莱索托': '非洲', '坦桑尼亚': '非洲', '肯尼亚': '非洲', '加蓬': '非洲',
    '索马里': '非洲', '纳米比亚': '非洲', '阿尔及利亚': '非洲', '南苏丹': '非洲',
    '博茨瓦纳': '非洲', '圣多美和普林西比': '非洲',
    # 东南亚
    '越南': '东南亚', '菲律宾': '东南亚', '马来西亚': '东南亚', '新加坡': '东南亚',
    '柬埔寨': '东南亚', '泰国': '东南亚', '印度尼西亚': '东南亚', '缅甸': '东南亚',
    '老挝': '东南亚', '文莱': '东南亚', '东帝汶': '东南亚',
    # 东亚
    '中国': '东亚', '日本': '东亚', '韩国': '东亚', '香港': '东亚',
    '台湾': '东亚', '蒙古': '东亚',
    # 南亚
    '巴基斯坦': '南亚', '印度': '南亚', '斯里兰卡': '南亚', '孟加拉国': '南亚',
    '尼泊尔': '南亚', '马尔代夫': '南亚',
    # 欧洲
    '欧盟': '欧洲', '英国': '欧洲', '荷兰': '欧洲', '波兰': '欧洲',
    '罗马尼亚': '欧洲', '法国': '欧洲', '比利时': '欧洲', '丹麦': '欧洲',
    '葡萄牙': '欧洲', '西班牙': '欧洲', '德国': '欧洲', '意大利': '欧洲',
    '白俄罗斯': '欧洲', '阿塞拜疆': '欧洲', '格鲁吉亚': '欧洲',
    '匈牙利': '欧洲', '捷克': '欧洲', '斯洛伐克': '欧洲', '奥地利': '欧洲',
    '瑞典': '欧洲', '挪威': '欧洲', '芬兰': '欧洲', '希腊': '欧洲',
    '乌克兰': '欧洲', '爱尔兰': '欧洲', '瑞士': '欧洲',
    # 北美洲
    '美国': '北美洲', '墨西哥': '北美洲', '加拿大': '北美洲',
    '多米尼加': '北美洲', '多米尼克': '北美洲', '多米尼加共和国': '北美洲',
    '哥斯达黎加': '北美洲', '巴拿马': '北美洲', '尼加拉瓜': '北美洲',
    '危地马拉': '北美洲', '洪都拉斯': '北美洲', '萨尔瓦多': '北美洲',
    '古巴': '北美洲', '牙买加': '北美洲', '海地': '北美洲',
    # 南美洲
    '巴西': '南美洲', '哥伦比亚': '南美洲', '秘鲁': '南美洲',
    '乌拉圭': '南美洲', '智利': '南美洲', '玻利维亚': '南美洲',
    '阿根廷': '南美洲', '厄瓜多尔': '南美洲', '委内瑞拉': '南美洲',
    # 独联体-中亚
    '乌兹别克斯坦': '独联体-中亚', '哈萨克斯坦': '独联体-中亚',
    '俄罗斯': '独联体-中亚', '吉尔吉斯斯坦': '独联体-中亚',
    '塔吉克斯坦': '独联体-中亚', '土库曼斯坦': '独联体-中亚',
    '亚美尼亚': '独联体-中亚', '摩尔多瓦': '独联体-中亚',
    # 大洋洲
    '澳大利亚': '大洋洲', '新西兰': '大洋洲',
}

# ============================================================
# 手动修正：基于序号（全部企业工作表中的序号）
# ============================================================
MANUAL_COUNTRY_FIXES_BY_SEQ = {
    # Damaco 系列：seq 228 总部比利时，229 土耳其，230 阿联酋，231 巴西，232 越南
    228: '比利时',
    230: '阿联酋',
    231: '巴西',
    232: '越南',
    # 中国香港 -> 香港
    1429: '香港',
    # 国家为"-"的企业，根据名称和简介判断
    1727: '安哥拉',   # INALCA - ANGOLA
    1728: '安哥拉',   # MERCADO FRESCO DE ANGOLA
    1729: '安哥拉',   # D E W DIAMOND VENTURES (安哥拉)
    1730: '安哥拉',   # AGROPEC - ANGOLA
    1731: '安哥拉',   # 质数商品 PRIME NUMBERS COMMODITIES
    1732: '安哥拉',   # ANGOALISSAR
    1733: '乌兹别克斯坦',  # ООО HILOLFOOD (俄文公司名，HILOL为乌兹别克语)
    1734: '安哥拉',   # COMANDO DA 2 REGIAO MILITAR (葡语，安哥拉军区)
    1735: '巴西',     # COMANDO DA MARINHA DO RIO DE JANEIRO
    1736: '哥伦比亚', # CENTRO INTERNACIONAL DE INVERSIONES S
    1737: '安哥拉',   # 第十二军区司令部
    1738: '巴西',     # 圣玛丽亚补给站
    1739: '安哥拉',   # IMOIND - ANGOLA
    1740: '哥伦比亚', # SENORA DEL CARMEN Y AMIGOS S A S
    1741: '美国',     # HANGE LLC
    1742: '刚果（金）',  # ETS MONGA BANZA
    1743: '安哥拉',   # IMOIND COMERCIO GERAL
    1744: '俄罗斯',   # ООО TRADEWAVE
    1745: '阿联酋',   # ARVIJAH FOODS TRADE LLC
    # 错误/AI生成内容
    1746: '伊朗',     # Pars Poultry Equipment (Pars=波斯/伊朗)
    1747: '伊朗',     # Behin Protein Robat (波斯语名称)
    1748: '伊拉克',   # Al-Raed Plastic Co. (阿拉伯语名)
    # 待核实企业，根据简介判断
    1754: '待核实',   # Mira Poultry & Feed - 信息不足
    1755: '沙特阿拉伯',  # Al-Fakieh Poultry Farm - 简介明确提到沙特
    1756: '阿联酋',   # Al-Rawabi Dairy Co. - 简介明确提到阿联酋
    1757: '阿联酋',   # Al-Kabeer Group - 简介明确提到阿联酋
    1758: '待核实',   # 未知公司
    1759: '待核实',   # Mohammad Janan Trading - 信息不足
    1760: '待核实',   # Arslan Poultry Ltd
    1761: '待核实',   # Dar Poultry Ltd
    1762: '待核实',   # 混合饲料
    1763: '待核实',   # 佩特罗斯农场
    1764: '待核实',   # 丹达尼农场
    1765: '待核实',   # 鸡肉价格大涨 (非企业)
    1766: '待核实',   # 新鲜肉类市场
    1767: '多米尼加共和国',  # GANADERA DOÑA MECHO SRL - 简介明确提到多米尼加共和国
    2314: '阿联酋',   # El-Eman Food Trading - 简介明确提到迪拜
}

def get_continent(country):
    """根据国家获取大洲"""
    if country in ('待核实', '-', ''):
        return '其他'
    return COUNTRY_CONTINENT.get(country, '其他')

def normalize_bilingual_text(text):
    """
    规范化双语文本：英文在前，中文在后
    统一格式为: EN: ...\n\nCN: ...
    """
    if not text or not text.strip():
        return text or ''
    
    text = text.strip()
    
    # 检测是否包含 EN: / CN: 标记
    has_en_marker = bool(re.search(r'\bEN\s*:', text, re.IGNORECASE))
    has_cn_marker = bool(re.search(r'\bCN\s*:', text, re.IGNORECASE))
    
    if has_en_marker or has_cn_marker:
        # 提取 EN 部分
        en_match = re.search(r'EN\s*:\s*(.*?)(?=\n\s*CN\s*:|\Z)', text, re.DOTALL | re.IGNORECASE)
        # 提取 CN 部分
        cn_match = re.search(r'CN\s*:\s*(.*?)(?=\n\s*EN\s*:|\Z)', text, re.DOTALL | re.IGNORECASE)
        
        en_text = en_match.group(1).strip() if en_match else ''
        cn_text = cn_match.group(1).strip() if cn_match else ''
        
        # 清理嵌套的 EN:/CN: 标记
        en_text = re.sub(r'^(EN|CN)\s*:\s*', '', en_text, flags=re.IGNORECASE).strip()
        cn_text = re.sub(r'^(EN|CN)\s*:\s*', '', cn_text, flags=re.IGNORECASE).strip()
        
        if en_text and cn_text:
            return f"EN: {en_text}\n\nCN: {cn_text}"
        elif en_text:
            return f"EN: {en_text}"
        elif cn_text:
            return f"CN: {cn_text}"
        else:
            return text
    
    # 无标记，检测语言
    has_chinese = bool(re.search(r'[\u4e00-\u9fff]', text))
    has_english = bool(re.search(r'[a-zA-Z]{3,}', text))
    
    if not has_chinese or not has_english:
        return text
    
    # 有换行分隔的中英文段落
    if '\n' in text:
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        cn_lines = []
        en_lines = []
        for line in lines:
            cn_count = len(re.findall(r'[\u4e00-\u9fff]', line))
            en_count = len(re.findall(r'[a-zA-Z]', line))
            if cn_count > 0 and en_count == 0:
                cn_lines.append(line)
            elif en_count > 0 and cn_count == 0:
                en_lines.append(line)
            elif cn_count >= en_count * 0.3:
                cn_lines.append(line)
            else:
                en_lines.append(line)
        
        if en_lines and cn_lines:
            return f"EN: {' '.join(en_lines)}\n\nCN: {' '.join(cn_lines)}"
    
    return text

def normalize_company_name(name):
    """
    规范化公司名称：英文在前，中文在后
    """
    if not name or not name.strip():
        return name or ''
    
    name = name.strip()
    has_chinese = bool(re.search(r'[\u4e00-\u9fff]', name))
    has_english = bool(re.search(r'[a-zA-Z]{2,}', name))
    
    if not has_chinese or not has_english:
        return name
    
    # 换行分隔的情况
    if '\n' in name:
        lines = [l.strip() for l in name.split('\n') if l.strip()]
        cn_lines = []
        en_lines = []
        for line in lines:
            cn_count = len(re.findall(r'[\u4e00-\u9fff]', line))
            en_count = len(re.findall(r'[a-zA-Z]', line))
            if cn_count > 0 and en_count == 0:
                cn_lines.append(line)
            elif en_count > 0 and cn_count == 0:
                en_lines.append(line)
            elif cn_count > en_count:
                cn_lines.append(line)
            else:
                en_lines.append(line)
        
        if en_lines and cn_lines:
            return '\n'.join(en_lines + cn_lines)
        return name
    
    # 单行：中文开头 + 括号英文 -> 英文开头 + 括号中文
    # 如: "中文名 (English Name)" -> "English Name (中文名)"
    m = re.match(r'^([\u4e00-\u9fff\s\-··（）()、，,]+?)\s*[（(]([A-Za-z][^）)]{2,})[）)](.*)$', name)
    if m:
        cn_part = m.group(1).strip()
        en_part = m.group(2).strip()
        rest = m.group(3).strip()
        if rest:
            return f"{en_part} ({cn_part}) {rest}"
        return f"{en_part} ({cn_part})"
    
    # 英文开头 + 括号中文 -> 已正确
    m2 = re.match(r'^([A-Za-z][^（(]*?)\s*[（(]([\u4e00-\u9fff][^）)]*)[）)](.*)$', name)
    if m2:
        return name
    
    return name

def normalize_short_field(text):
    """
    规范化短字段（核心角色、采购倾向）：英文在前，中文在后
    """
    if not text or not text.strip():
        return text or ''
    
    text = text.strip()
    has_chinese = bool(re.search(r'[\u4e00-\u9fff]', text))
    has_english = bool(re.search(r'[a-zA-Z]{2,}', text))
    
    if not has_chinese or not has_english:
        return text
    
    # 格式: "中文描述 (English Description)" -> "English Description (中文描述)"
    m = re.match(r'^([\u4e00-\u9fff\s\-·/、，,·]+?)\s*[（(]([A-Za-z][^）)]{2,})[）)](.*)$', text)
    if m:
        cn_part = m.group(1).strip()
        en_part = m.group(2).strip()
        rest = m.group(3).strip()
        if rest:
            return f"{en_part} ({cn_part}) {rest}"
        return f"{en_part} ({cn_part})"
    
    # 英文在前括号中文 -> 已正确
    m2 = re.match(r'^([A-Za-z][^（(]*?)\s*[（(]([\u4e00-\u9fff][^）)]*)[）)](.*)$', text)
    if m2:
        return text
    
    # 换行分隔
    if '\n' in text:
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        cn_lines = [l for l in lines if re.search(r'[\u4e00-\u9fff]', l) and not re.search(r'[a-zA-Z]{3,}', l)]
        en_lines = [l for l in lines if re.search(r'[a-zA-Z]{3,}', l) and not re.search(r'[\u4e00-\u9fff]', l)]
        if en_lines and cn_lines:
            return ' / '.join(en_lines) + ' (' + ' / '.join(cn_lines) + ')'
    
    return text

def process_all_companies():
    """处理所有企业数据"""
    print("读取原始 Excel 文件...")
    wb = load_workbook(INPUT_FILE, read_only=True)
    ws_all = wb['全部企业']
    
    companies = []
    
    for row in ws_all.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        
        seq_no = row[0]
        seq_int = int(seq_no) if seq_no and str(seq_no).isdigit() else None
        
        orig_name = str(row[1] or '').strip()
        orig_country = str(row[2] or '').strip()
        orig_role = str(row[3] or '').strip() if row[3] else ''
        orig_tendency = str(row[4] or '').strip() if row[4] else ''
        orig_profile = str(row[5] or '').strip() if row[5] else ''
        orig_products = str(row[6] or '').strip() if row[6] else ''
        website = str(row[7] or '').strip() if row[7] else ''
        contact = str(row[8] or '').strip() if row[8] else ''
        china = str(row[9] or '否').strip() if row[9] else '否'
        
        # 1. 修正国家
        if seq_int and seq_int in MANUAL_COUNTRY_FIXES_BY_SEQ:
            fixed_country = MANUAL_COUNTRY_FIXES_BY_SEQ[seq_int]
        elif len(orig_country) > 30 or orig_country in ('Error', '无法确定', '无法识别', '北非/中东', '-'):
            # 异常国家字段，根据简介判断（默认归其他）
            fixed_country = '待核实'
        elif orig_country == '中国香港':
            fixed_country = '香港'
        else:
            fixed_country = orig_country
        
        continent = get_continent(fixed_country)
        
        # 2. 规范化文本字段
        fixed_name = normalize_company_name(orig_name)
        fixed_role = normalize_short_field(orig_role)
        fixed_tendency = normalize_short_field(orig_tendency)
        fixed_profile = normalize_bilingual_text(orig_profile)
        fixed_products = normalize_bilingual_text(orig_products)
        
        companies.append({
            'seq': seq_no,
            'seq_int': seq_int,
            'name': fixed_name,
            'orig_name': orig_name,
            'country': fixed_country,
            'orig_country': orig_country,
            'continent': continent,
            'role': fixed_role,
            'tendency': fixed_tendency,
            'profile': fixed_profile,
            'products': fixed_products,
            'website': website,
            'contact': contact,
            'china': china,
            'country_changed': fixed_country != orig_country,
        })
    
    print(f"处理完成，共 {len(companies)} 家企业")
    
    # 统计修改情况
    country_changed = [c for c in companies if c['country_changed']]
    print(f"国家/地区修正: {len(country_changed)} 家")
    for c in country_changed:
        print(f"  [{c['seq']}] {c['orig_name'][:45]} | {c['orig_country'][:20]} -> {c['country']}")
    
    return companies

def write_excel(companies):
    """写入新的 Excel 文件"""
    print("\n生成新 Excel 文件...")
    
    wb_new = openpyxl.Workbook()
    
    # 样式定义
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    continent_colors = {
        '中东': 'FFF2CC', '非洲': 'E2EFDA', '东南亚': 'DDEBF7',
        '东亚': 'FCE4D6', '南亚': 'F4CCFF', '欧洲': 'D9E1F2',
        '北美洲': 'FFE699', '南美洲': 'C6EFCE', '独联体-中亚': 'F8CBAD',
        '大洋洲': 'BDD7EE', '其他': 'F2F2F2',
    }
    
    headers = [
        '序号', '公司名称\n(英文在前，中文在后)', '国家/地区', '大洲',
        '核心角色', '采购/合作倾向',
        '公司简介\n(英文在前，中文在后)',
        '主营产品\n(英文在前，中文在后)',
        '网站及社交媒体', '邮箱/电话', '是否在中国采购过'
    ]
    col_widths = [7, 42, 14, 12, 32, 36, 65, 42, 30, 30, 12]
    
    continents_order = ['中东', '非洲', '东南亚', '东亚', '南亚', '欧洲', '北美洲', '南美洲', '独联体-中亚', '大洋洲', '其他']
    
    def setup_sheet(ws, title_text=None):
        ws.row_dimensions[1].height = 40
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'
    
    def write_company_row(ws, row_idx, company):
        continent = company['continent']
        fill_color = continent_colors.get(continent, 'FFFFFF')
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        values = [
            company['seq'], company['name'], company['country'], company['continent'],
            company['role'], company['tendency'], company['profile'],
            company['products'], company['website'], company['contact'], company['china'],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 60
    
    # ---- 全部企业工作表 ----
    ws_all = wb_new.active
    ws_all.title = '全部企业'
    setup_sheet(ws_all)
    for row_idx, company in enumerate(companies, 2):
        write_company_row(ws_all, row_idx, company)
    print(f"  [全部企业] {len(companies)} 家企业")
    
    # ---- 各大洲分工作表 ----
    for continent_name in continents_order:
        continent_companies = [c for c in companies if c['continent'] == continent_name]
        if not continent_companies:
            continue
        ws_c = wb_new.create_sheet(title=continent_name)
        setup_sheet(ws_c)
        for row_idx, company in enumerate(continent_companies, 2):
            write_company_row(ws_c, row_idx, company)
        print(f"  [{continent_name}] {len(continent_companies)} 家企业")
    
    # ---- 修改记录工作表 ----
    ws_log = wb_new.create_sheet(title='修改记录')
    log_headers = ['序号', '公司名称（原）', '原国家', '修正后国家', '修正后大洲', '修改原因']
    log_widths = [7, 50, 20, 15, 12, 60]
    log_header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws_log.row_dimensions[1].height = 35
    for col_idx, (header, width) in enumerate(zip(log_headers, log_widths), 1):
        cell = ws_log.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = log_header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws_log.column_dimensions[get_column_letter(col_idx)].width = width
    ws_log.freeze_panes = 'A2'
    
    changed_companies = [c for c in companies if c['country_changed']]
    for row_idx, company in enumerate(changed_companies, 2):
        orig = company['orig_country']
        fixed = company['country']
        seq_int = company['seq_int']
        
        if orig == '-':
            reason = '原国家字段为空（"-"），根据公司名称和简介判断所属国家'
        elif orig in ('Error', '无法确定', '无法识别', '北非/中东') or len(orig) > 30:
            reason = '原国家字段为AI生成错误内容，已根据简介修正'
        elif orig == '中国香港':
            reason = '"中国香港"统一规范为"香港"'
        elif orig == '土耳其' and fixed in ('比利时', '阿联酋', '巴西', '越南'):
            reason = f'Damaco集团分支机构，实际注册地为{fixed}，原误标为土耳其'
        else:
            reason = f'根据公司简介和名称，原国家"{orig}"有误，修正为"{fixed}"'
        
        values = [company['seq'], company['orig_name'][:60], orig, fixed, company['continent'], reason]
        for col_idx, value in enumerate(values, 1):
            cell = ws_log.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        ws_log.row_dimensions[row_idx].height = 35
    
    print(f"  [修改记录] {len(changed_companies)} 条")
    
    # ---- 统计摘要工作表 ----
    ws_stat = wb_new.create_sheet(title='统计摘要')
    ws_stat.column_dimensions['A'].width = 20
    ws_stat.column_dimensions['B'].width = 12
    
    ws_stat['A1'] = '全球禽肉进口商数据库 - 统计摘要（修正版 v2.0）'
    ws_stat['A1'].font = Font(name='Arial', bold=True, size=14)
    ws_stat.merge_cells('A1:B1')
    ws_stat.row_dimensions[1].height = 30
    
    ws_stat['A3'] = '大洲'
    ws_stat['B3'] = '企业数量'
    for cell in [ws_stat['A3'], ws_stat['B3']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    continent_stats = {}
    for c in companies:
        continent_stats[c['continent']] = continent_stats.get(c['continent'], 0) + 1
    
    row = 4
    for continent_name in continents_order:
        count = continent_stats.get(continent_name, 0)
        ws_stat.cell(row=row, column=1, value=continent_name).border = thin_border
        ws_stat.cell(row=row, column=2, value=count).border = thin_border
        row += 1
    
    total_cell_a = ws_stat.cell(row=row, column=1, value='合计')
    total_cell_b = ws_stat.cell(row=row, column=2, value=len(companies))
    total_cell_a.font = Font(bold=True)
    total_cell_b.font = Font(bold=True)
    total_cell_a.border = thin_border
    total_cell_b.border = thin_border
    
    ws_stat.cell(row=row+2, column=1, value='国家/地区修正数量').border = thin_border
    ws_stat.cell(row=row+2, column=2, value=len(changed_companies)).border = thin_border
    ws_stat.cell(row=row+3, column=1, value='数据版本').border = thin_border
    ws_stat.cell(row=row+3, column=2, value='修正版 v2.0').border = thin_border
    
    wb_new.save(OUTPUT_FILE)
    print(f"\n✅ 新 Excel 文件已保存: {OUTPUT_FILE}")
    return companies

if __name__ == '__main__':
    companies = process_all_companies()
    write_excel(companies)
    
    # 输出 JSON 供数据库导入
    json_output = []
    for c in companies:
        json_output.append({
            'seqNo': c['seq'],
            'companyName': c['name'],
            'country': c['country'],
            'continent': c['continent'],
            'coreRole': c['role'],
            'purchaseTendency': c['tendency'],
            'companyProfile': c['profile'],
            'mainProducts': c['products'],
            'websiteSocial': c['website'],
            'contactInfo': c['contact'],
            'hasPurchasedFromChina': c['china'],
        })
    
    with open('/home/ubuntu/companies_fixed.json', 'w', encoding='utf-8') as f:
        json.dump(json_output, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON 数据已保存: /home/ubuntu/companies_fixed.json")
